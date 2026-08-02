"""
Excel 指标录入解析：读取单元格数值与底色来源。
"""
from __future__ import annotations

from io import BytesIO
from math import isnan
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook

from apps.coredata.excel_color_sources import build_cell_payload
from apps.coredata.management.commands.indicator_zh_en import (
    AREA_INDIMAP,
    AREA_INDIMAP_UNIT,
    INDIMAP,
    INDIMAP_UNIT,
)

# Excel 表头第二行经 _ 拼入列名；匹配指标时只取 _ 前第一段，单位以 *_UNIT map 为准。

# Excel 列名简称 → AREA_INDIMAP 中的中文名（比对前预处理，目标名须在 map 内）
_AREA_EXCEL_NAME_ALIASES = {
    "常住人口数": "常住人口",
    "户籍人口数": "户籍人口",
    "一般预算收入": "一般公共预算收入",
    "人均一般预算收入": "人均一般公共预算收入",
    "城镇居民家庭人均可支配收入": "城镇居民人均可支配收入",
    "农村居民家庭人均纯收入": "农村居民人均可支配收入",
    "农民居民人均纯收入": "农村居民人均可支配收入",
}


def match_city_indicator_name(name_zh: str) -> Optional[Tuple[str, str]]:
    """在地市统一映射 INDIMAP / INDIMAP_UNIT 中查找。"""
    text = (name_zh or "").strip()
    if not text:
        return None
    name_en = INDIMAP.get(text)
    if not name_en or name_en not in INDIMAP_UNIT:
        return None
    canonical = INDIMAP_UNIT[name_en].get("name_zh") or text
    return canonical, name_en


def match_area_indicator_name(name_zh: str) -> Optional[Tuple[str, str]]:
    """
    在区县统一映射 AREA_INDIMAP / AREA_INDIMAP_UNIT 中查找。
    命中返回 (规范中文名, name_en)；未命中返回 None。
    """
    text = (name_zh or "").strip()
    if not text:
        return None
    name_en = AREA_INDIMAP.get(text)
    if not name_en or name_en not in AREA_INDIMAP_UNIT:
        return None
    canonical = AREA_INDIMAP_UNIT[name_en].get("name_zh") or text
    return canonical, name_en


def strip_unit_suffix(name: str) -> str:
    import re
    text = (name or "").strip().lstrip("#")
    return re.sub(r"\([^)]*\)$", "", text).strip()


def excel_column_name_base(col_name: str) -> str:
    """Excel 列名按 _ 拆分，取第一部分作为指标名（后面通常是单位，不参与匹配）。"""
    raw = str(col_name).split("__dup", 1)[0].strip().replace("\n", "").replace("\r", "")
    raw = strip_unit_suffix(raw)
    if "_" in raw:
        return raw.split("_", 1)[0].strip()
    return raw


def resolve_city_excel_indicator(
    col_name: str,
    last_metric_raw: Optional[str] = None,
) -> Tuple[Optional[str], Optional[str]]:
    """地市 Excel 列名 → 取 _ 前第一段 → 与 INDIMAP 比对，规范名取自 INDIMAP_UNIT。"""
    raw = str(col_name).split("__dup", 1)[0].strip().replace("\n", "").replace("\r", "")
    if raw in ("城市", "A") or raw.startswith("Column_"):
        return None, None

    base = excel_column_name_base(col_name)
    if base == "增长率":
        prev_base = excel_column_name_base(last_metric_raw) if last_metric_raw else ""
        matched = match_city_indicator_name(f"{prev_base}增长率")
        if matched:
            return matched
        return match_city_indicator_name("GDP增长率") or (None, None)

    for candidate in (base, raw):
        matched = match_city_indicator_name(candidate)
        if matched:
            return matched
    return None, None


def resolve_area_excel_indicator(
    col_name: str,
    last_metric_raw: Optional[str] = None,
) -> Tuple[Optional[str], Optional[str]]:
    """
    区县 Excel 列名 → 取 _ 前第一段 → 与 AREA_INDIMAP 比对。
    命中则返回 (规范中文名, name_en)，单位以 AREA_INDIMAP_UNIT 为准。
    """
    raw = str(col_name).split("__dup", 1)[0].strip().replace("\n", "").replace("\r", "")
    if raw in ("所辖区域名称", "区域名称", "省份名称", "城市名称", "城市", "地市"):
        return None, None

    base = excel_column_name_base(col_name)

    if base == "增长率":
        prev_base = excel_column_name_base(last_metric_raw) if last_metric_raw else ""
        prev_base = _AREA_EXCEL_NAME_ALIASES.get(prev_base, prev_base)
        matched = match_area_indicator_name(f"{prev_base}增长率")
        if matched:
            return matched
        return match_area_indicator_name("GDP增长率") or (None, None)

    candidate = _AREA_EXCEL_NAME_ALIASES.get(base, base)
    return match_area_indicator_name(candidate) or (None, None)


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _build_two_level_headers(ws, header_row_count: int = 2) -> List[str]:
    headers: List[str] = []
    max_col = ws.max_column or 0
    for col_idx in range(1, max_col + 1):
        parts = []
        for row_idx in range(1, header_row_count + 1):
            text = _cell_text(ws.cell(row=row_idx, column=col_idx).value)
            if text:
                parts.append(text)
        if parts:
            headers.append("_".join(parts))
        else:
            headers.append(f"Column_{col_idx}")
    return headers


def _build_flexible_headers(ws) -> Tuple[List[str], int]:
    """区县表：自动识别 1 或 2 行表头。"""
    max_col = ws.max_column or 0

    def _is_data_row(row_idx: int) -> bool:
        values = []
        for col_idx in range(1, max_col + 1):
            text = _cell_text(ws.cell(row=row_idx, column=col_idx).value)
            if text:
                values.append(text)
        if not values:
            return False
        numeric_like = 0
        for v in values:
            try:
                float(v.replace(",", ""))
                numeric_like += 1
            except (TypeError, ValueError):
                pass
        return numeric_like >= max(2, len(values) // 3)

    header_rows = 1
    if ws.max_row and ws.max_row > 1 and not _is_data_row(2):
        header_rows = 2

    headers: List[str] = []
    for col_idx in range(1, max_col + 1):
        level1 = _cell_text(ws.cell(row=1, column=col_idx).value)
        level2 = _cell_text(ws.cell(row=2, column=col_idx).value) if header_rows == 2 else ""
        if header_rows == 2 and level1 and level2 and level1 != level2:
            col_name = f"{level1}_{level2}"
        else:
            col_name = level1 or level2 or f"Column_{col_idx}"
        headers.append(col_name.replace("\n", "").replace("\r", "").strip())

    seen: Dict[str, int] = {}
    unique_headers: List[str] = []
    for name in headers:
        seen[name] = seen.get(name, 0) + 1
        if seen[name] == 1:
            unique_headers.append(name)
        else:
            unique_headers.append(f"{name}__dup{seen[name]}")

    return unique_headers, header_rows


def _row_has_data(ws, row_idx: int, max_col: int) -> bool:
    for col_idx in range(1, max_col + 1):
        if _cell_text(ws.cell(row=row_idx, column=col_idx).value):
            return True
    return False


def parse_city_indicator_excel(file_obj) -> List[Dict[str, Any]]:
    """
    地市指标录入表：第 1-2 行为表头，第 4 行起为数据（与现有 upload_excel 一致）。
    返回每行 dict，指标列为 {value, source, note}。
    """
    content = file_obj.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    headers = _build_two_level_headers(ws, header_row_count=2)
    data_start_row = 4
    max_col = len(headers)
    rows: List[Dict[str, Any]] = []

    for row_idx in range(data_start_row, (ws.max_row or 0) + 1):
        if not _row_has_data(ws, row_idx, max_col):
            continue
        row_data: Dict[str, Any] = {}
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if col_name in ("城市", "A") or col_name.startswith("Column_"):
                row_data[col_name] = cell.value
            else:
                row_data[col_name] = build_cell_payload(cell)
        rows.append(row_data)

    return rows


def parse_area_indicator_excel(file_obj) -> List[Dict[str, Any]]:
    """区县指标录入表：自动识别表头行数，数据行带底色来源。"""
    content = file_obj.read()
    wb = load_workbook(filename=BytesIO(content), data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb.active

    headers, header_rows = _build_flexible_headers(ws)
    data_start_row = header_rows + 1
    max_col = len(headers)
    rows: List[Dict[str, Any]] = []

    for row_idx in range(data_start_row, (ws.max_row or 0) + 1):
        if not _row_has_data(ws, row_idx, max_col):
            continue
        row_data: Dict[str, Any] = {}
        for col_idx, col_name in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            skip_meta = col_name in (
                "城市", "城市名称", "地市", "area", "所辖区县名称",
                "所辖区域名称", "区县", "区县名称", "区县名", "区域名称", "A",
            ) or col_name.startswith("Column_")
            if skip_meta:
                row_data[col_name] = cell.value
            else:
                row_data[col_name] = build_cell_payload(cell)
        rows.append(row_data)

    return rows


def extract_cell_fields(raw_value: Any) -> Tuple[Any, str, str]:
    """兼容普通值与 {value, source, note} 结构。"""
    if isinstance(raw_value, dict) and "value" in raw_value:
        return (
            raw_value.get("value"),
            raw_value.get("source") or "",
            raw_value.get("note") or "",
        )
    return raw_value, "", ""


def has_excel_cell_value(value: Any) -> bool:
    """Excel 单元格无有效数值时不应写入数据库。"""
    if value is None:
        return False
    if isinstance(value, float) and isnan(value):
        return False
    if isinstance(value, str):
        text = value.strip()
        if not text:
            return False
        if text.lower() in ("nan", "none"):
            return False
    return True
